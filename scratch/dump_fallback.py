import openpyxl

def dump_fallback_books(filepath, output_txt):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    if 'FallbackBooks' in wb.sheetnames:
        sheet = wb['FallbackBooks']
        rows = list(sheet.iter_rows(values_only=True))
        with open(output_txt, 'w', encoding='utf-8') as f:
            headers = rows[0]
            f.write(f"{headers}\n")
            for r in rows[1:]:
                f.write(f"BookId: {r[0]}, Lang: {r[1]}, PageIdx: {r[2]}, Side: {r[3]}, Type: {r[4]}, Title: {r[5]}, Context: {r[6]}, Author: {r[7]}, Text: {r[8]}, Verse: {r[9]}\n")

dump_fallback_books('data.xlsx', 'inspect_books_output.txt')
print("Done dumping fallback books")
