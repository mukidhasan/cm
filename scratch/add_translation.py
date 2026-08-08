import openpyxl

wb = openpyxl.load_workbook('data.xlsx')
if 'Translations' in wb.sheetnames:
    sheet = wb['Translations']
    # Check if 'hero_name' is already there
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == 'hero_name':
            row[1].value = 'ukid Hasan Seiam'
            row[2].value = 'মুকিদ হাসান সিয়াম'
            found = True
            break
    
    if not found:
        sheet.append(['hero_name', 'ukid Hasan Seiam', 'মুকিদ হাসান সিয়াম'])
    
    wb.save('data.xlsx')
    print('Successfully added/updated hero_name in Translations sheet.')
else:
    print('Translations sheet not found!')
